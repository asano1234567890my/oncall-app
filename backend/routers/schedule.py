from __future__ import annotations

import calendar as _calendar
import datetime
import io
import urllib.parse
from typing import List, Optional
from uuid import UUID

import uuid

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from core.auth import get_current_hospital
from core.db import get_db
from models.doctor import Doctor
from models.hospital import Hospital
from models.shift import ShiftAssignment
from services.settings_service import (
    get_draft_schedule,
    get_published_months_by_doctor_token,
    upsert_draft_schedule,
    delete_draft_schedule,
)

from fastapi.responses import Response
from sqlalchemy.orm import selectinload

router = APIRouter(prefix="/api/schedule", tags=["Schedule"])

DAY_SHIFT_LABEL = chr(0x65E5) + chr(0x76F4)
NIGHT_SHIFT_LABEL = chr(0x5F53) + chr(0x76F4)


class ShiftData(BaseModel):
    day: int
    day_shift: Optional[UUID] = None
    night_shift: Optional[UUID] = None


class SaveScheduleRequest(BaseModel):
    year: int
    month: int
    num_doctors: int
    schedule: List[ShiftData]


class SaveDraftRequest(BaseModel):
    schedule: List[ShiftData]


class RangeShiftItem(BaseModel):
    date: str
    shift_type: str
    doctor_id: str


def _month_bounds(year: int, month: int) -> tuple[datetime.date, datetime.date]:
    start_date = datetime.date(year, month, 1)
    if month == 12:
        end_date = datetime.date(year + 1, 1, 1)
    else:
        end_date = datetime.date(year, month + 1, 1)
    return start_date, end_date


def _normalize_shift_type(raw_shift_type: str) -> str:
    if raw_shift_type in {DAY_SHIFT_LABEL, "day", "day_shift"}:
        return "day"
    if raw_shift_type in {NIGHT_SHIFT_LABEL, "night", "night_shift"}:
        return "night"
    return str(raw_shift_type)


## ── ICS Feed (Google Calendar sync) ──


def _build_ics(doctor_name: str, hospital_name: str, assignments: list, doctor_id: str = "") -> str:
    """Build an iCalendar (.ics) string from shift assignments."""
    now_stamp = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//ShifRaku//oncall-app//JP",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:{doctor_name} 当直表",
    ]

    for a in assignments:
        shift_label = "日直" if _normalize_shift_type(a.shift_type) == "day" else "当直"
        d = a.date
        next_day = d + datetime.timedelta(days=1)
        dtstart = f"{d.year:04d}{d.month:02d}{d.day:02d}"
        dtend = f"{next_day.year:04d}{next_day.month:02d}{next_day.day:02d}"

        # UID must be ASCII — use doctor_id (UUID) instead of doctor_name
        uid = f"{d.isoformat()}-{_normalize_shift_type(a.shift_type)}-{doctor_id or 'doc'}@shiftraku.com"
        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{now_stamp}",
            f"DTSTART;VALUE=DATE:{dtstart}",
            f"DTEND;VALUE=DATE:{dtend}",
            f"SUMMARY:{shift_label}（{hospital_name}）",
            f"DESCRIPTION:{doctor_name} {shift_label}",
            "STATUS:CONFIRMED",
            "END:VEVENT",
        ])

    lines.append("END:VCALENDAR")
    return "\r\n".join(lines)


@router.get("/ical/{doctor_token}")
async def get_ical_feed(
    doctor_token: str,
    year: Optional[int] = Query(None, description="対象年（指定時はその月のみ）"),
    month: Optional[int] = Query(None, description="対象月（yearと併用）"),
    db: AsyncSession = Depends(get_db),
):
    """ICSフィード — 医師個人の確定済みシフトを .ics 形式で返す（終日イベント）。公開月のみ。認証不要。"""
    result = await db.execute(
        select(Doctor)
        .options(selectinload(Doctor.hospital))
        .where(Doctor.access_token == doctor_token)
    )
    doctor = result.scalar_one_or_none()
    if doctor is None:
        raise HTTPException(status_code=404, detail="Invalid token")

    published = await get_published_months_by_doctor_token(db, doctor_token)
    published_set = set(published)

    # 年月指定時はその月のみ、未指定時は過去3ヶ月〜未来6ヶ月
    if year and month:
        start, end = _month_bounds(year, month)
    else:
        today = datetime.date.today()
        start = today.replace(day=1) - datetime.timedelta(days=90)
        end = today + datetime.timedelta(days=180)

    shift_result = await db.execute(
        select(ShiftAssignment)
        .where(
            ShiftAssignment.doctor_id == doctor.id,
            ShiftAssignment.date >= start,
            ShiftAssignment.date <= end,
        )
        .order_by(ShiftAssignment.date)
    )
    all_assignments = shift_result.scalars().all()
    assignments = [
        a for a in all_assignments
        if f"{a.date.year}-{a.date.month:02d}" in published_set
    ]

    hospital_name = doctor.hospital.name if doctor.hospital else "病院"
    ics_content = _build_ics(doctor.name, hospital_name, assignments, doctor_id=str(doctor.id))

    return Response(
        content=ics_content,
        media_type="text/calendar; charset=utf-8",
        headers={
            "Content-Disposition": "attachment; filename=\"shifts.ics\"; filename*=UTF-8''" + urllib.parse.quote(f"{doctor.name}_shifts.ics"),
            "Cache-Control": "no-cache, no-store, must-revalidate",
        },
    )


@router.get("/public/{doctor_token}/{year}/{month}")
async def get_public_schedule(
    doctor_token: str,
    year: int,
    month: int,
    db: AsyncSession = Depends(get_db),
):
    """トークンで認証し、公開月の全体スケジュールを医師名付きで返す。"""
    from services.settings_service import get_published_months, get_system_setting

    result = await db.execute(
        select(Doctor)
        .options(selectinload(Doctor.hospital))
        .where(Doctor.access_token == doctor_token)
    )
    doctor = result.scalar_one_or_none()
    if doctor is None:
        raise HTTPException(status_code=404, detail="Invalid token")

    month_key = f"{year}-{month:02d}"
    published = await get_published_months(db, doctor.hospital_id)
    if month_key not in published:
        return {"published": False, "schedule": [], "doctors": {}}

    # 全医師名マップ
    doctors_result = await db.execute(
        select(Doctor).where(Doctor.hospital_id == doctor.hospital_id)
    )
    all_doctors = doctors_result.scalars().all()
    doctor_name_map = {str(d.id): d.name for d in all_doctors}

    start_date, end_date = _month_bounds(year, month)
    hospital_doctor_ids = [d.id for d in all_doctors]

    shift_result = await db.execute(
        select(ShiftAssignment)
        .where(
            ShiftAssignment.date >= start_date,
            ShiftAssignment.date < end_date,
            ShiftAssignment.doctor_id.in_(hospital_doctor_ids),
        )
        .order_by(ShiftAssignment.date)
    )
    assignments = shift_result.scalars().all()

    days_in_month = _calendar.monthrange(year, month)[1]
    formatted: dict = {
        day: {"day": day, "day_shift": None, "night_shift": None}
        for day in range(1, days_in_month + 1)
    }
    for a in assignments:
        day = a.date.day
        name = doctor_name_map.get(str(a.doctor_id), "?")
        if _normalize_shift_type(a.shift_type) == "day":
            formatted[day]["day_shift"] = name
        else:
            formatted[day]["night_shift"] = name

    # Fetch publish comment for this month
    publish_comment = None
    raw = await get_system_setting(db, doctor.hospital_id, f"publish_comment_{month_key}")
    if raw and isinstance(raw, str):
        publish_comment = raw

    return {
        "published": True,
        "schedule": sorted(formatted.values(), key=lambda x: x["day"]),
        "doctors": doctor_name_map,
        "publish_comment": publish_comment,
    }


@router.get("/public-shifts/{doctor_token}")
async def get_public_shifts(
    doctor_token: str,
    db: AsyncSession = Depends(get_db),
):
    """医師トークンからその医師の確定済みシフトをJSON返却。公開月のみ。認証不要。"""
    result = await db.execute(
        select(Doctor).where(Doctor.access_token == doctor_token)
    )
    doctor = result.scalar_one_or_none()
    if doctor is None:
        raise HTTPException(status_code=404, detail="Invalid token")

    published = await get_published_months_by_doctor_token(db, doctor_token)
    published_set = set(published)

    today = datetime.date.today()
    start = today.replace(day=1) - datetime.timedelta(days=90)
    end = today + datetime.timedelta(days=180)

    shift_result = await db.execute(
        select(ShiftAssignment)
        .where(
            ShiftAssignment.doctor_id == doctor.id,
            ShiftAssignment.date >= start,
            ShiftAssignment.date <= end,
        )
        .order_by(ShiftAssignment.date)
    )
    assignments = shift_result.scalars().all()

    return [
        {
            "date": str(a.date),
            "shift_type": a.shift_type,
        }
        for a in assignments
        if f"{a.date.year}-{a.date.month:02d}" in published_set
    ]


@router.post("/save")
async def save_schedule(
    req: SaveScheduleRequest,
    hospital_id: uuid.UUID = Depends(get_current_hospital),
    db: AsyncSession = Depends(get_db),
):
    try:
        assigned_doctor_ids = {
            doctor_id
            for item in req.schedule
            for doctor_id in (item.day_shift, item.night_shift)
            if doctor_id is not None
        }

        if assigned_doctor_ids:
            result = await db.execute(
                select(Doctor.id).where(
                    Doctor.id.in_(assigned_doctor_ids),
                    Doctor.hospital_id == hospital_id,
                )
            )
            existing_doctor_ids = set(result.scalars().all())
            missing_doctor_ids = assigned_doctor_ids - existing_doctor_ids
            if missing_doctor_ids:
                missing_values = ", ".join(
                    sorted(str(doctor_id) for doctor_id in missing_doctor_ids)
                )
                raise HTTPException(
                    status_code=400,
                    detail=f"Unknown doctor_id in schedule: {missing_values}",
                )

        start_date, end_date = _month_bounds(req.year, req.month)

        # hospital_id経由でフィルタ（shift_assignmentsはdoctorに紐づく）
        hospital_doctor_ids = (
            await db.execute(
                select(Doctor.id).where(Doctor.hospital_id == hospital_id)
            )
        ).scalars().all()

        await db.execute(
            delete(ShiftAssignment).where(
                ShiftAssignment.date >= start_date,
                ShiftAssignment.date < end_date,
                ShiftAssignment.doctor_id.in_(hospital_doctor_ids),
            )
        )

        new_assignments = []
        for item in req.schedule:
            current_date = datetime.date(req.year, req.month, item.day)

            if item.day_shift is not None:
                new_assignments.append(
                    ShiftAssignment(
                        date=current_date,
                        doctor_id=item.day_shift,
                        shift_type=DAY_SHIFT_LABEL,
                    )
                )

            if item.night_shift is not None:
                new_assignments.append(
                    ShiftAssignment(
                        date=current_date,
                        doctor_id=item.night_shift,
                        shift_type=NIGHT_SHIFT_LABEL,
                    )
                )

        db.add_all(new_assignments)
        await db.commit()

        return {
            "success": True,
            "message": "\u30b7\u30d5\u30c8\u3092\u30c7\u30fc\u30bf\u30d9\u30fc\u30b9\u306b\u4fdd\u5b58\u3057\u307e\u3057\u305f\uff01",
            "saved_count": len(new_assignments),
        }

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        print(f"DEBUG Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/range", response_model=List[RangeShiftItem])
async def get_schedule_range(
    start_date: datetime.date,
    end_date: datetime.date,
    hospital_id: uuid.UUID = Depends(get_current_hospital),
    db: AsyncSession = Depends(get_db),
):
    if start_date > end_date:
        raise HTTPException(
            status_code=400,
            detail="start_date must be on or before end_date",
        )

    try:
        hospital_doctor_ids = (
            await db.execute(
                select(Doctor.id).where(Doctor.hospital_id == hospital_id)
            )
        ).scalars().all()

        result = await db.execute(
            select(ShiftAssignment)
            .where(
                ShiftAssignment.date >= start_date,
                ShiftAssignment.date <= end_date,
                ShiftAssignment.doctor_id.in_(hospital_doctor_ids),
            )
            .order_by(
                ShiftAssignment.date,
                ShiftAssignment.shift_type,
                ShiftAssignment.doctor_id,
            )
        )
        assignments = result.scalars().all()

        return [
            {
                "date": assignment.date.isoformat(),
                "shift_type": _normalize_shift_type(assignment.shift_type),
                "doctor_id": str(assignment.doctor_id),
            }
            for assignment in assignments
        ]

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{year}/{month}")
async def delete_schedule(
    year: int,
    month: int,
    hospital_id: uuid.UUID = Depends(get_current_hospital),
    db: AsyncSession = Depends(get_db),
):
    """管理者用：指定月のシフトをDBから完全削除する。フロントエンドには非公開。"""
    try:
        start_date, end_date = _month_bounds(year, month)
        hospital_doctor_ids = (
            await db.execute(
                select(Doctor.id).where(Doctor.hospital_id == hospital_id)
            )
        ).scalars().all()
        await db.execute(
            delete(ShiftAssignment).where(
                ShiftAssignment.date >= start_date,
                ShiftAssignment.date < end_date,
                ShiftAssignment.doctor_id.in_(hospital_doctor_ids),
            )
        )
        await db.commit()
        return {"success": True, "message": f"{year}年{month}月のシフトを削除しました"}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


## ── Draft Schedule ──


@router.get("/draft/{year}/{month}")
async def get_draft(
    year: int,
    month: int,
    hospital_id: uuid.UUID = Depends(get_current_hospital),
    db: AsyncSession = Depends(get_db),
):
    data = await get_draft_schedule(db, hospital_id, year, month)
    if data is None:
        return {"schedule": None, "saved_at": None}
    return {"schedule": data.get("schedule"), "saved_at": data.get("saved_at")}


@router.put("/draft/{year}/{month}")
async def save_draft(
    year: int,
    month: int,
    req: SaveDraftRequest,
    hospital_id: uuid.UUID = Depends(get_current_hospital),
    db: AsyncSession = Depends(get_db),
):
    schedule_data = [
        {"day": item.day, "day_shift": str(item.day_shift) if item.day_shift else None, "night_shift": str(item.night_shift) if item.night_shift else None}
        for item in req.schedule
    ]
    saved_at = await upsert_draft_schedule(db, hospital_id, year, month, schedule_data)
    return {"success": True, "saved_at": saved_at}


@router.delete("/draft/{year}/{month}")
async def remove_draft(
    year: int,
    month: int,
    hospital_id: uuid.UUID = Depends(get_current_hospital),
    db: AsyncSession = Depends(get_db),
):
    await delete_draft_schedule(db, hospital_id, year, month)
    return {"success": True}


## ── Stats (年間集計) ──


@router.get("/stats")
async def get_schedule_stats(
    year: int = Query(...),
    hospital_id: uuid.UUID = Depends(get_current_hospital),
    db: AsyncSession = Depends(get_db),
):
    """年間のシフトデータを集計して返す。"""
    import jpholiday

    # 医師一覧
    doctors_result = await db.execute(
        select(Doctor)
        .where(Doctor.hospital_id == hospital_id, Doctor.is_active == True)  # noqa: E712
        .order_by(Doctor.name)
    )
    doctors = doctors_result.scalars().all()
    doctor_map = {str(d.id): d.name for d in doctors}

    # 年間シフト（1月〜12月）
    start_date = datetime.date(year, 1, 1)
    end_date = datetime.date(year, 12, 31)

    doctor_ids = [d.id for d in doctors]
    if not doctor_ids:
        return {"doctors": [], "shifts": [], "holidays": []}

    shifts_result = await db.execute(
        select(ShiftAssignment)
        .where(
            ShiftAssignment.date >= start_date,
            ShiftAssignment.date <= end_date,
            ShiftAssignment.doctor_id.in_(doctor_ids),
        )
        .order_by(ShiftAssignment.date)
    )
    shifts = shifts_result.scalars().all()

    # 祝日一覧
    holiday_dates = set()
    for m in range(1, 13):
        days_in_month = _calendar.monthrange(year, m)[1]
        for d in range(1, days_in_month + 1):
            dt = datetime.date(year, m, d)
            if jpholiday.is_holiday(dt):
                holiday_dates.add(dt.isoformat())

    return {
        "doctors": [{"id": str(d.id), "name": d.name} for d in doctors],
        "shifts": [
            {
                "date": s.date.isoformat(),
                "doctor_id": str(s.doctor_id),
                "shift_type": _normalize_shift_type(s.shift_type),
            }
            for s in shifts
        ],
        "holidays": sorted(holiday_dates),
    }


## ── Export (PDF / Excel) ──


WEEKDAY_LABELS = ["月", "火", "水", "木", "金", "土", "日"]


def _weekday_ja(year: int, month: int, day: int) -> str:
    return WEEKDAY_LABELS[datetime.date(year, month, day).weekday()]


async def _fetch_export_data(
    year: int, month: int, hospital_id: uuid.UUID, db: AsyncSession
):
    """Fetch hospital name, doctor map, schedule rows and holiday info for export."""
    hospital = await db.get(Hospital, hospital_id)
    hospital_name = hospital.name if hospital else "病院"

    doctors_result = await db.execute(
        select(Doctor).where(Doctor.hospital_id == hospital_id)
    )
    doctor_map = {d.id: d.name for d in doctors_result.scalars().all()}

    start_date, end_date = _month_bounds(year, month)
    hospital_doctor_ids = list(doctor_map.keys())

    result = await db.execute(
        select(ShiftAssignment)
        .where(
            ShiftAssignment.date >= start_date,
            ShiftAssignment.date < end_date,
            ShiftAssignment.doctor_id.in_(hospital_doctor_ids) if hospital_doctor_ids else ShiftAssignment.doctor_id.is_(None),
        )
        .order_by(ShiftAssignment.date)
    )
    assignments = result.scalars().all()

    days_in_month = _calendar.monthrange(year, month)[1]
    schedule: dict = {
        day: {"day": day, "day_shift": None, "night_shift": None}
        for day in range(1, days_in_month + 1)
    }
    for a in assignments:
        day = a.date.day
        if _normalize_shift_type(a.shift_type) == "day":
            schedule[day]["day_shift"] = str(a.doctor_id)
        else:
            schedule[day]["night_shift"] = str(a.doctor_id)

    rows = sorted(schedule.values(), key=lambda r: r["day"])

    # Determine holidays (weekday-based only; custom holidays require settings)
    try:
        import jpholiday
        holiday_dates = {
            h[0] for h in jpholiday.between(
                datetime.date(year, month, 1),
                datetime.date(year, month, days_in_month),
            )
        }
    except Exception:
        holiday_dates = set()

    return hospital_name, doctor_map, rows, holiday_dates


def _doctor_label(doctor_id: str | None, doctor_map: dict) -> str:
    if not doctor_id:
        return ""
    uid = uuid.UUID(doctor_id) if isinstance(doctor_id, str) else doctor_id
    return doctor_map.get(uid, "")


def _build_pdf(
    year: int, month: int, hospital_name: str, doctor_map: dict,
    rows: list, holiday_dates: set,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
    from reportlab.lib.styles import ParagraphStyle

    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    font_name = "HeiseiKakuGo-W5"  # ゴシック太め

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )

    elements = []

    # Title
    title_style = ParagraphStyle(
        "Title", fontName=font_name, fontSize=16, leading=22, alignment=1,
    )
    elements.append(Paragraph(f"{hospital_name}　{year}年{month}月 当直表", title_style))
    elements.append(Spacer(1, 6 * mm))

    # Split into 2 columns: left=1-15, right=16-end (Feb: 14/14-15)
    mid = 14 if len(rows) <= 28 else 15
    left_rows = rows[:mid]
    right_rows = rows[mid:]

    def make_column_data(row_list):
        header = ["日付", "日直", "当直"]
        data = [header]
        for r in row_list:
            d = r["day"]
            wd = _weekday_ja(year, month, d)
            date_label = f"{d}({wd})"
            dt = datetime.date(year, month, d)
            is_sun = dt.weekday() == 6
            is_sat = dt.weekday() == 5
            is_holiday = is_sun or dt in holiday_dates
            show_day = is_holiday or is_sat
            day_name = _doctor_label(r["day_shift"], doctor_map) if show_day else ""
            night_name = _doctor_label(r["night_shift"], doctor_map)
            data.append([date_label, day_name, night_name])
        return data

    left_data = make_column_data(left_rows)
    right_data = make_column_data(right_rows)

    # Pad shorter column
    while len(left_data) < len(right_data):
        left_data.append(["", "", ""])
    while len(right_data) < len(left_data):
        right_data.append(["", "", ""])

    # Combine into single table: [日付, 日直, 当直, gap, 日付, 日直, 当直]
    combined = []
    for i in range(len(left_data)):
        lr = left_data[i]
        rr = right_data[i]
        combined.append(lr + [""] + rr)

    page_w = A4[0] - 30 * mm  # available width
    page_h = A4[1] - 30 * mm  # available height
    col_date = 22 * mm
    col_shift = (page_w - 2 * col_date - 3 * mm) / 4  # 4 shift columns share remaining
    col_gap = 3 * mm
    col_widths = [col_date, col_shift, col_shift, col_gap, col_date, col_shift, col_shift]

    # Calculate row height to fill the page vertically
    # 31日 → 16+1ヘッダー=17行が最大。余裕を持って計算
    title_space = 22 + 8 * mm  # title + spacer
    bottom_margin_extra = 5 * mm  # 下部に余裕
    available_for_table = page_h - title_space - bottom_margin_extra
    max_rows = 17  # 31日の月: 右列16行+1ヘッダー=17行
    num_rows = len(combined)
    row_height = available_for_table / max(num_rows, max_rows)
    row_heights = [row_height] * num_rows

    table = Table(combined, colWidths=col_widths, rowHeights=row_heights, repeatRows=1)

    # Font size: fit to row height (roughly 60% of row height, capped)
    font_size = min(max(int(row_height * 0.55), 8), 14)
    title_font_size = font_size + 2

    # Style
    style_cmds = [
        # Font — ゴシック体（太め）
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        # Header row
        ("FONTSIZE", (0, 0), (-1, 0), font_size),
        ("BACKGROUND", (0, 0), (2, 0), colors.Color(0.93, 0.93, 0.93)),
        ("BACKGROUND", (4, 0), (6, 0), colors.Color(0.93, 0.93, 0.93)),
        ("FONTNAME", (0, 0), (-1, 0), font_name),
        # Alignment
        ("ALIGN", (1, 0), (2, -1), "CENTER"),
        ("ALIGN", (5, 0), (6, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Padding
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        # Bold outer borders for left table
        ("BOX", (0, 0), (2, -1), 1.5, colors.black),
        ("LINEBELOW", (0, 0), (2, 0), 1.5, colors.black),
        ("INNERGRID", (0, 0), (2, -1), 0.5, colors.Color(0.7, 0.7, 0.7)),
        # Bold outer borders for right table
        ("BOX", (4, 0), (6, -1), 1.5, colors.black),
        ("LINEBELOW", (4, 0), (6, 0), 1.5, colors.black),
        ("INNERGRID", (4, 0), (6, -1), 0.5, colors.Color(0.7, 0.7, 0.7)),
        # Gap column invisible
        ("LINEAFTER", (2, 0), (2, -1), 0, colors.white),
        ("LINEBEFORE", (4, 0), (4, -1), 0, colors.white),
    ]

    # Color rows for holidays/saturdays
    for col_offset, row_list in [(0, left_rows), (4, right_rows)]:
        for idx, r in enumerate(row_list):
            row_idx = idx + 1  # +1 for header
            dt = datetime.date(year, month, r["day"])
            if dt.weekday() == 6 or dt in holiday_dates:
                style_cmds.append(("TEXTCOLOR", (col_offset, row_idx), (col_offset, row_idx), colors.Color(0.85, 0.15, 0.15)))
                style_cmds.append(("BACKGROUND", (col_offset, row_idx), (col_offset + 2, row_idx), colors.Color(1, 0.95, 0.95)))
            elif dt.weekday() == 5:
                style_cmds.append(("TEXTCOLOR", (col_offset, row_idx), (col_offset, row_idx), colors.Color(0.15, 0.15, 0.85)))
                style_cmds.append(("BACKGROUND", (col_offset, row_idx), (col_offset + 2, row_idx), colors.Color(0.95, 0.95, 1)))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    return buf.getvalue()


def _build_xlsx(
    year: int, month: int, hospital_name: str, doctor_map: dict,
    rows: list, holiday_dates: set,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月"

    # ── Styles ──
    thin = Side(style="thin", color="999999")
    thick = Side(style="medium", color="000000")
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    sun_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    sat_fill = PatternFill(start_color="F0F0FF", end_color="F0F0FF", fill_type="solid")
    sun_font = Font(color="CC2222", size=10)
    sat_font = Font(color="2222CC", size=10)
    normal_font = Font(size=10)
    header_font = Font(bold=True, size=10)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # ── 左側: スケジュール（A〜C） ──
    # A=日付, B=日直, C=当直, D=gap, E以降=集計
    # L,M=非表示ヘルパー列（土曜○/日祝○ — COUNTIFSで参照）
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 2  # gap

    # Hidden helper columns for COUNTIFS
    ws.column_dimensions["L"].width = 5
    ws.column_dimensions["L"].hidden = True
    ws.column_dimensions["M"].width = 5
    ws.column_dimensions["M"].hidden = True

    sched_cols = 3  # A〜C

    # Title (spans schedule area)
    ws.merge_cells("A1:C1")
    ws["A1"] = f"{hospital_name}　{year}年{month}月 当直表"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # Schedule headers
    sched_headers = ["日付", "日直", "当直"]
    s_row = 3
    for col_idx, label in enumerate(sched_headers, 1):
        cell = ws.cell(row=s_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = Border(top=thick, bottom=thick, left=thick if col_idx == 1 else thin, right=thick if col_idx == sched_cols else thin)

    # Schedule data
    for idx, r in enumerate(rows):
        excel_row = s_row + 1 + idx
        d = r["day"]
        wd = _weekday_ja(year, month, d)
        dt = datetime.date(year, month, d)
        is_sun = dt.weekday() == 6
        is_sat = dt.weekday() == 5
        is_holiday = is_sun or dt in holiday_dates
        show_day = is_holiday or is_sat

        date_cell = ws.cell(row=excel_row, column=1, value=f"{d}({wd})")
        day_cell = ws.cell(row=excel_row, column=2, value=_doctor_label(r["day_shift"], doctor_map) if show_day else "")
        night_cell = ws.cell(row=excel_row, column=3, value=_doctor_label(r["night_shift"], doctor_map))

        # Hidden helper columns (L=土曜, M=日祝)
        ws.cell(row=excel_row, column=12, value="○" if is_sat else "")
        ws.cell(row=excel_row, column=13, value="○" if is_holiday else "")

        fill = sun_fill if is_holiday else sat_fill if is_sat else None
        font = sun_font if is_holiday else sat_font if is_sat else normal_font

        date_cell.alignment = left_align
        for c in (day_cell, night_cell):
            c.alignment = center

        for c in (date_cell, day_cell, night_cell):
            c.font = font
            if fill:
                c.fill = fill
            c.border = Border(
                top=thin, bottom=thin,
                left=thick if c.column == 1 else thin,
                right=thick if c.column == sched_cols else thin,
            )

    # Schedule bottom border
    last_sched_row = s_row + len(rows)
    for col_idx in range(1, sched_cols + 1):
        cell = ws.cell(row=last_sched_row, column=col_idx)
        b = cell.border
        cell.border = Border(top=b.top, bottom=thick, left=b.left, right=b.right)

    # ── 右側: 医師別集計（E〜J） ──
    sum_col_start = 5  # E列

    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 8

    # Summary title
    ws.merge_cells(start_row=1, start_column=sum_col_start, end_row=1, end_column=sum_col_start + 5)
    title_cell = ws.cell(row=1, column=sum_col_start, value="医師別集計")
    title_cell.font = title_font
    title_cell.alignment = center

    # Summary headers
    sum_headers = ["医師名", "当直回数", "日直回数", "土曜当直", "日祝当直", "合計"]
    for i, label in enumerate(sum_headers):
        col = sum_col_start + i
        cell = ws.cell(row=s_row, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = Border(top=thick, bottom=thick, left=thick if i == 0 else thin, right=thick if i == len(sum_headers) - 1 else thin)

    # Collect unique doctors
    doctor_names = sorted({
        _doctor_label(r[shift], doctor_map)
        for r in rows
        for shift in ("day_shift", "night_shift")
        if r[shift] and _doctor_label(r[shift], doctor_map)
    })

    # COUNTIF ranges (B=日直, C=当直, L=土曜helper, M=日祝helper)
    days_count = len(rows)
    data_start = s_row + 1
    data_end = s_row + days_count
    night_range = f"$C${data_start}:$C${data_end}"
    day_range = f"$B${data_start}:$B${data_end}"
    sat_range = f"$L${data_start}:$L${data_end}"
    sunhol_range = f"$M${data_start}:$M${data_end}"

    from openpyxl.utils import get_column_letter
    name_col_letter = get_column_letter(sum_col_start)  # H

    for i, name in enumerate(doctor_names):
        r = s_row + 1 + i
        name_ref = f"${name_col_letter}${r}"

        ws.cell(row=r, column=sum_col_start, value=name).font = normal_font
        ws.cell(row=r, column=sum_col_start).alignment = left_align

        ws.cell(row=r, column=sum_col_start + 1).value = f'=COUNTIF({night_range},{name_ref})'
        ws.cell(row=r, column=sum_col_start + 1).font = normal_font
        ws.cell(row=r, column=sum_col_start + 1).alignment = center

        ws.cell(row=r, column=sum_col_start + 2).value = f'=COUNTIF({day_range},{name_ref})'
        ws.cell(row=r, column=sum_col_start + 2).font = normal_font
        ws.cell(row=r, column=sum_col_start + 2).alignment = center

        ws.cell(row=r, column=sum_col_start + 3).value = f'=COUNTIFS({night_range},{name_ref},{sat_range},"○")'
        ws.cell(row=r, column=sum_col_start + 3).font = normal_font
        ws.cell(row=r, column=sum_col_start + 3).alignment = center

        ws.cell(row=r, column=sum_col_start + 4).value = f'=COUNTIFS({night_range},{name_ref},{sunhol_range},"○")'
        ws.cell(row=r, column=sum_col_start + 4).font = normal_font
        ws.cell(row=r, column=sum_col_start + 4).alignment = center

        i_col = get_column_letter(sum_col_start + 1)
        j_col = get_column_letter(sum_col_start + 2)
        ws.cell(row=r, column=sum_col_start + 5).value = f'={i_col}{r}+{j_col}{r}'
        ws.cell(row=r, column=sum_col_start + 5).font = Font(bold=True, size=10)
        ws.cell(row=r, column=sum_col_start + 5).alignment = center

        for j in range(len(sum_headers)):
            col = sum_col_start + j
            ws.cell(row=r, column=col).border = Border(
                top=thin, bottom=thin,
                left=thick if j == 0 else thin,
                right=thick if j == len(sum_headers) - 1 else thin,
            )

    # Summary bottom border
    last_sum_row = s_row + len(doctor_names)
    for j in range(len(sum_headers)):
        col = sum_col_start + j
        cell = ws.cell(row=last_sum_row, column=col)
        b = cell.border
        cell.border = Border(top=b.top, bottom=thick, left=b.left, right=b.right)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_xlsx_simple(
    year: int, month: int, hospital_name: str, doctor_map: dict,
    rows: list, holiday_dates: set,
) -> bytes:
    """統計なしのシンプルな当直表Excel（医師向け公開エクスポート用）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月"

    thin = Side(style="thin", color="999999")
    thick = Side(style="medium", color="000000")
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    sun_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    sat_fill = PatternFill(start_color="F0F0FF", end_color="F0F0FF", fill_type="solid")
    sun_font = Font(color="CC2222", size=10)
    sat_font = Font(color="2222CC", size=10)
    normal_font = Font(size=10)
    header_font = Font(bold=True, size=10)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # A列は余白、データはB列(2)から開始
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    # タイトル: B2
    ws.merge_cells("B2:D2")
    ws["B2"] = f"{hospital_name}　{year}年{month}月 当直表"
    ws["B2"].font = title_font
    ws["B2"].alignment = center

    # ヘッダー: 4行目、B〜D列 (col 2〜4)
    s_row = 4
    col_start = 2  # B列
    for col_offset, label in enumerate(["日付", "日直", "当直"]):
        col_idx = col_start + col_offset
        cell = ws.cell(row=s_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = Border(top=thick, bottom=thick, left=thick if col_offset == 0 else thin, right=thick if col_offset == 2 else thin)

    for idx, r in enumerate(rows):
        excel_row = s_row + 1 + idx
        d = r["day"]
        wd = _weekday_ja(year, month, d)
        dt = datetime.date(year, month, d)
        is_sun = dt.weekday() == 6
        is_sat = dt.weekday() == 5
        is_holiday = is_sun or dt in holiday_dates
        show_day = is_holiday or is_sat

        date_cell = ws.cell(row=excel_row, column=col_start, value=f"{d}({wd})")
        day_cell = ws.cell(row=excel_row, column=col_start + 1, value=_doctor_label(r["day_shift"], doctor_map) if show_day else "")
        night_cell = ws.cell(row=excel_row, column=col_start + 2, value=_doctor_label(r["night_shift"], doctor_map))

        fill = sun_fill if is_holiday else sat_fill if is_sat else None
        font = sun_font if is_holiday else sat_font if is_sat else normal_font

        date_cell.alignment = left_align
        for c in (day_cell, night_cell):
            c.alignment = center

        for c_idx, c in enumerate((date_cell, day_cell, night_cell)):
            c.font = font
            if fill:
                c.fill = fill
            c.border = Border(
                top=thin, bottom=thin,
                left=thick if c_idx == 0 else thin,
                right=thick if c_idx == 2 else thin,
            )

    last_row = s_row + len(rows)
    for col_offset in range(3):
        cell = ws.cell(row=last_row, column=col_start + col_offset)
        b = cell.border
        cell.border = Border(top=b.top, bottom=thick, left=b.left, right=b.right)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/public-export/{doctor_token}/{year}/{month}")
async def public_export_schedule(
    doctor_token: str,
    year: int,
    month: int,
    format: str = Query("pdf", pattern="^(pdf|xlsx)$"),
    db: AsyncSession = Depends(get_db),
):
    """トークン認証で当直表をPDF/Excelダウンロード（統計なし）。"""
    from services.settings_service import get_published_months

    result = await db.execute(
        select(Doctor)
        .options(selectinload(Doctor.hospital))
        .where(Doctor.access_token == doctor_token)
    )
    doctor = result.scalar_one_or_none()
    if doctor is None:
        raise HTTPException(status_code=404, detail="Invalid token")

    month_key = f"{year}-{month:02d}"
    published = await get_published_months(db, doctor.hospital_id)
    if month_key not in published:
        raise HTTPException(status_code=403, detail="この月は公開されていません。")

    hospital_name, doctor_map, rows, holiday_dates = await _fetch_export_data(
        year, month, doctor.hospital_id, db,
    )

    if not any(r["day_shift"] or r["night_shift"] for r in rows):
        raise HTTPException(status_code=404, detail="この月の当直表はまだ保存されていません。")

    from urllib.parse import quote
    filename_ja = f"{hospital_name}_{year}年{month}月_当直表"
    filename_ascii = f"oncall_{year}_{month:02d}"

    if format == "xlsx":
        content = _build_xlsx_simple(year, month, hospital_name, doctor_map, rows, holiday_dates)
        ext = "xlsx"
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        content = _build_pdf(year, month, hospital_name, doctor_map, rows, holiday_dates)
        ext = "pdf"
        media = "application/pdf"

    return Response(
        content=content,
        media_type=media,
        headers={
            "Content-Disposition": f"attachment; filename=\"{filename_ascii}.{ext}\"; filename*=UTF-8''{quote(filename_ja)}.{ext}",
        },
    )


@router.get("/export/{year}/{month}")
async def export_schedule(
    year: int,
    month: int,
    format: str = Query("pdf", pattern="^(pdf|xlsx)$"),
    hospital_id: uuid.UUID = Depends(get_current_hospital),
    db: AsyncSession = Depends(get_db),
):
    """当直表をPDFまたはExcelファイルとしてダウンロード。"""
    hospital_name, doctor_map, rows, holiday_dates = await _fetch_export_data(
        year, month, hospital_id, db,
    )

    if not any(r["day_shift"] or r["night_shift"] for r in rows):
        raise HTTPException(status_code=404, detail="この月の当直表はまだ保存されていません。")

    from urllib.parse import quote
    filename_ja = f"{hospital_name}_{year}年{month}月_当直表"
    filename_ascii = f"oncall_{year}_{month:02d}"

    if format == "xlsx":
        content = _build_xlsx(year, month, hospital_name, doctor_map, rows, holiday_dates)
        ext = "xlsx"
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        content = _build_pdf(year, month, hospital_name, doctor_map, rows, holiday_dates)
        ext = "pdf"
        media = "application/pdf"

    return Response(
        content=content,
        media_type=media,
        headers={
            "Content-Disposition": f"attachment; filename=\"{filename_ascii}.{ext}\"; filename*=UTF-8''{quote(filename_ja)}.{ext}",
        },
    )


## ── GET /{year}/{month} — must be LAST (catch-all path pattern) ──


@router.get("/{year}/{month}")
async def get_schedule(
    year: int,
    month: int,
    hospital_id: uuid.UUID = Depends(get_current_hospital),
    db: AsyncSession = Depends(get_db),
):
    try:
        start_date, end_date = _month_bounds(year, month)
        hospital_doctor_ids = (
            await db.execute(
                select(Doctor.id).where(Doctor.hospital_id == hospital_id)
            )
        ).scalars().all()

        result = await db.execute(
            select(ShiftAssignment)
            .where(
                ShiftAssignment.date >= start_date,
                ShiftAssignment.date < end_date,
                ShiftAssignment.doctor_id.in_(hospital_doctor_ids),
            )
            .order_by(ShiftAssignment.date)
        )
        assignments = result.scalars().all()

        days_in_month = _calendar.monthrange(year, month)[1]
        formatted_data: dict = {
            day: {"day": day, "day_shift": None, "night_shift": None}
            for day in range(1, days_in_month + 1)
        }

        for assignment in assignments:
            day = assignment.date.day
            if _normalize_shift_type(assignment.shift_type) == "day":
                formatted_data[day]["day_shift"] = str(assignment.doctor_id)
            else:
                formatted_data[day]["night_shift"] = str(assignment.doctor_id)

        return sorted(formatted_data.values(), key=lambda item: item["day"])

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

